"""
report_generator.py - Professional report generation for MEPI results

Generates:
  - Summary statistics tables
  - Key findings and insights
  - Regional analysis tables
  - Dimension analysis tables
  - Formatted Excel workbooks (with colour coding)
  - PDF-style text reports

Usage:
    from report_generator import ReportGenerator
    rg = ReportGenerator(results_df)
    rg.export_excel_report("output/mepi_report.xlsx")
    rg.export_text_report("output/mepi_report.txt")
"""

import os
from datetime import datetime

import numpy as np
import pandas as pd

from config import DIMENSIONS, DISTRICT_COLUMN, DIVISION_COLUMN, POVERTY_LABELS
from statistical_analysis import StatisticalAnalyzer
from spatial_analysis import SpatialAnalyzer


class ReportGenerator:
    """
    Generate professional MEPI reports in multiple formats.

    Parameters
    ----------
    results_df : pd.DataFrame
        Output of ``MEPICalculator.calculate()``.
    title : str, optional
        Report title.  Default: ``"MEPI Analysis Report – Bangladesh"``.
    """

    def __init__(
        self,
        results_df: pd.DataFrame,
        title: str = "MEPI Analysis Report – Bangladesh",
    ):
        self.df = results_df.copy()
        self.title = title
        self.generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        self._stat = StatisticalAnalyzer(self.df)
        self._spatial = SpatialAnalyzer(self.df)
        self._name_col = (
            "upazila_name" if "upazila_name" in self.df.columns else self.df.columns[0]
        )
        self._dim_score_cols = [
            f"{d}_score" for d in DIMENSIONS if f"{d}_score" in self.df.columns
        ]

    # ------------------------------------------------------------------
    # Internal data builders
    # ------------------------------------------------------------------

    def _summary_stats(self) -> pd.DataFrame:
        return self._stat.descriptive_statistics()

    def _category_table(self) -> pd.DataFrame:
        if "poverty_category" not in self.df.columns:
            return pd.DataFrame()
        counts = self.df["poverty_category"].value_counts().rename("Count")
        pct = (counts / len(self.df) * 100).round(1).rename("Share (%)")
        return pd.concat([counts, pct], axis=1).reset_index().rename(
            columns={"index": "Poverty Category"}
        )

    def _regional_table(self, group_col: str) -> pd.DataFrame:
        if group_col not in self.df.columns:
            return pd.DataFrame()
        score_cols = ["mepi_score"] + self._dim_score_cols
        tbl = (
            self.df.groupby(group_col)[score_cols]
            .mean()
            .round(4)
        )
        tbl["n_upazilas"] = self.df.groupby(group_col).size()
        return tbl.sort_values("mepi_score", ascending=False).reset_index()

    def _top_bottom_table(self, n: int = 10) -> tuple:
        top = self._spatial.top_n_upazilas(n)
        bottom = self._spatial.bottom_n_upazilas(n)
        return top, bottom

    def _dimension_contrib_table(self) -> pd.DataFrame:
        return self._stat.dimension_contribution()

    def _inequality_table(self) -> pd.DataFrame:
        return self._stat.inequality_summary()

    def _key_findings(self) -> list:
        """Generate a list of key finding strings from the data."""
        findings = []
        mean_mepi = self.df["mepi_score"].mean()
        findings.append(
            f"The mean MEPI score across {len(self.df)} upazilas is {mean_mepi:.3f}."
        )

        if "poverty_category" in self.df.columns:
            for cat in ["Severely Poor", "Moderately Poor", "Non-Poor"]:
                n = (self.df["poverty_category"] == cat).sum()
                pct = round(n / len(self.df) * 100, 1)
                findings.append(f"  {n} upazilas ({pct}%) are classified as '{cat}'.")

        # Most deprived dimension
        contrib = self._stat.dimension_contribution()
        if not contrib.empty:
            top_dim = contrib.iloc[0]
            findings.append(
                f"The '{top_dim['Dimension']}' dimension contributes most to energy poverty "
                f"({top_dim['Contribution Share (%)']:.1f}% of the weighted MEPI score)."
            )

        # Gini
        gini = self._stat.gini_coefficient()
        findings.append(
            f"The Gini coefficient of MEPI scores is {gini:.4f}, indicating "
            + ("low" if gini < 0.2 else "moderate" if gini < 0.4 else "high")
            + " inequality in energy poverty across upazilas."
        )

        # Most/least deprived
        worst = self.df.nlargest(1, "mepi_score").iloc[0]
        best = self.df.nsmallest(1, "mepi_score").iloc[0]
        findings.append(
            f"The most energy-poor upazila is '{worst[self._name_col]}' (MEPI={worst['mepi_score']:.3f})."
        )
        findings.append(
            f"The least energy-poor upazila is '{best[self._name_col]}' (MEPI={best['mepi_score']:.3f})."
        )

        return findings

    # ------------------------------------------------------------------
    # Text report
    # ------------------------------------------------------------------

    def export_text_report(self, filepath: str) -> str:
        """
        Write a plain-text summary report.

        Parameters
        ----------
        filepath : str
            Path for the output text file.

        Returns
        -------
        str
            Absolute path to the saved file.
        """
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

        lines = []
        sep = "=" * 70

        lines += [
            sep,
            self.title,
            f"Generated: {self.generated_at}",
            sep,
            "",
            "KEY FINDINGS",
            "-" * 40,
        ]
        for finding in self._key_findings():
            lines.append(f"• {finding}")

        lines += ["", "DESCRIPTIVE STATISTICS", "-" * 40]
        lines.append(self._summary_stats().to_string())

        lines += ["", "POVERTY CATEGORY BREAKDOWN", "-" * 40]
        cat_tbl = self._category_table()
        if not cat_tbl.empty:
            lines.append(cat_tbl.to_string(index=False))

        lines += ["", "DIMENSION CONTRIBUTION", "-" * 40]
        lines.append(self._dimension_contrib_table().to_string(index=False))

        lines += ["", "INEQUALITY MEASURES", "-" * 40]
        lines.append(self._inequality_table().to_string(index=False))

        if DIVISION_COLUMN in self.df.columns:
            lines += ["", "DIVISION-LEVEL SUMMARY", "-" * 40]
            lines.append(self._regional_table(DIVISION_COLUMN).to_string(index=False))

        top, bottom = self._top_bottom_table(10)
        show_cols = [c for c in ["rank", self._name_col, "district", "mepi_score", "poverty_category"] if c in top.columns]

        lines += ["", "TOP 10 MOST ENERGY-POOR UPAZILAS", "-" * 40]
        lines.append(top[show_cols].to_string(index=False))

        lines += ["", "BOTTOM 10 LEAST ENERGY-POOR UPAZILAS", "-" * 40]
        lines.append(bottom[show_cols].to_string(index=False))

        lines += ["", sep]

        content = "\n".join(lines)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

        print(f"Text report saved: {filepath}")
        return os.path.abspath(filepath)

    # ------------------------------------------------------------------
    # Excel report
    # ------------------------------------------------------------------

    def export_excel_report(
        self,
        filepath: str,
        include_charts: bool = False,
    ) -> str:
        """
        Write a formatted Excel workbook with multiple sheets.

        Sheets included:
          - **Overview**      : key findings and metadata
          - **All Results**   : full MEPI results table
          - **Summary Stats** : descriptive statistics
          - **By Division**   : division-level aggregation
          - **By District**   : district-level aggregation
          - **By Zone**       : geographic zone analysis
          - **Top 10 / Bottom 10** : extreme upazilas
          - **Dimension Contribution** : contribution analysis
          - **Inequality**    : Gini and Theil measures

        Parameters
        ----------
        filepath : str
            Path for the output ``.xlsx`` file.
        include_charts : bool, optional
            Reserved for future use.  Currently ignored.

        Returns
        -------
        str
            Absolute path to the saved file.
        """
        try:
            from openpyxl.styles import (
                PatternFill,
                Font,
                Alignment,
                Border,
                Side,
            )
            from openpyxl.utils import get_column_letter
            HAS_OPENPYXL_STYLES = True
        except ImportError:
            HAS_OPENPYXL_STYLES = False

        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

        top10, bottom10 = self._top_bottom_table(10)

        sheets: dict = {
            "All_Results": self.df,
            "Summary_Stats": self._summary_stats().reset_index().rename(columns={"index": "Score"}),
            "Poverty_Categories": self._category_table(),
            "Dimension_Contribution": self._dimension_contrib_table(),
            "Inequality": self._inequality_table(),
            "Top_10_Most_Poor": top10,
            "Bottom_10_Least_Poor": bottom10,
        }

        if DIVISION_COLUMN in self.df.columns:
            sheets["By_Division"] = self._regional_table(DIVISION_COLUMN)
        if DISTRICT_COLUMN in self.df.columns:
            sheets["By_District"] = self._regional_table(DISTRICT_COLUMN)
        if "geographic_zone" in self._spatial.df.columns:
            zone_tbl = self._spatial.zone_comparison()
            if not zone_tbl.empty:
                sheets["By_Zone"] = zone_tbl

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Overview sheet first
            overview_data = {
                "Item": ["Report Title", "Generated At", "Total Upazilas", "Mean MEPI Score", "Gini Coefficient"] + [""] + ["KEY FINDINGS"] + [f"  {i+1}." for i in range(len(self._key_findings()))],
                "Value": [self.title, self.generated_at, len(self.df), round(self.df["mepi_score"].mean(), 4), self._stat.gini_coefficient()] + [""] + [""] + self._key_findings(),
            }
            pd.DataFrame(overview_data).to_excel(writer, sheet_name="Overview", index=False)

            for sheet_name, df_sheet in sheets.items():
                if df_sheet is not None and not df_sheet.empty:
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

            if HAS_OPENPYXL_STYLES:
                _apply_excel_styling(writer)

        print(f"Excel report saved: {filepath}")
        return os.path.abspath(filepath)


# ---------------------------------------------------------------------------
# Excel styling helper
# ---------------------------------------------------------------------------

def _apply_excel_styling(writer):
    """Apply basic formatting to all sheets in the Excel workbook."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Poverty category colours
    cat_fills = {
        "Non-Poor": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Moderately Poor": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Severely Poor": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    wb = writer.book
    for ws in wb.worksheets:
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Alternate row shading
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    if cell.fill.fill_type == "none":
                        cell.fill = alt_fill

        # Colour-code poverty category column
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "poverty_category":
                for data_row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for data_cell in data_row:
                        if data_cell.value in cat_fills:
                            data_cell.fill = cat_fills[data_cell.value]
                break

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
